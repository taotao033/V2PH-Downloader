"""Local archival sync driver for V2PH-Downloader.

Two-step pipeline meant for **personal, offline archival only**:

    discover - Use the v2dl web bot to enumerate top-level listing URLs
               (companies, optionally actors) and write them to a watch
               list file.
    sync     - Read the watch list and invoke ``v2dl --input-file ...``
               for the actual download. Already-downloaded albums are
               skipped automatically via ``download_log_path``.

This script intentionally hardcodes polite rate-limit defaults
(``--max-worker 2``, ``--rate-limit 1000``) and a sleep between top-level
URLs. Do not crank them up - the source site is rate-limited per IP and
account, and being noisy will only get you (and the tool) banned faster.

Typical usage::

    # one-time discovery of all companies -> data/sync/companies.xlsx
    # (columns: url, name, total, 是否采集(0/1). Default 采集 = 0.)
    python scripts/sync_local.py discover companies

    # open the xlsx in Excel and set 是否采集 = 1 for the companies you
    # want to mirror locally, then save.

    # daily incremental sync (only re-checks the first page of each
    # listing, so already-downloaded albums get skipped instantly).
    # Only rows with 是否采集 = 1 in the xlsx will be synced.
    python scripts/sync_local.py sync --destination "D:/v2ph_archive" --mode incremental

    # weekly full re-scan (paginates through every listing page; still
    # skips albums that are already on disk).
    python scripts/sync_local.py sync --destination "D:/v2ph_archive" --mode full
"""

from __future__ import annotations

import re
import sys
import time
import asyncio
import argparse
import tempfile
import subprocess
from pathlib import Path
from typing import Any

from lxml import html

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DATA_DIR = REPO_ROOT / "data" / "sync"
# Companies now live in an xlsx so the user can flip a "是否采集" column
# per row in Excel without touching the script. We still recognise the
# legacy .txt path so existing setups don't break silently.
DEFAULT_COMPANIES_FILE = DEFAULT_DATA_DIR / "companies.xlsx"
LEGACY_COMPANIES_TXT = DEFAULT_DATA_DIR / "companies.txt"
DEFAULT_ACTORS_FILE = DEFAULT_DATA_DIR / "actors.xlsx"
LEGACY_ACTORS_TXT = DEFAULT_DATA_DIR / "actors.txt"
DEFAULT_WATCH_FILE = DEFAULT_DATA_DIR / "watch.txt"

# Excel column headers. Order matters - readers (sync) and writers
# (discover) both rely on it.
XLSX_HEADER = ["url", "name", "total", "是否采集(0/1)"]
# Index of the "是否采集" column for sync-time filtering.
XLSX_COLLECT_COL = 3

# Minimum sleep between top-level URLs (companies / actors). v2dl already
# delays between pagination requests within a single listing; this one
# adds extra slack between the listings themselves.
INTER_URL_SLEEP_SECONDS = 5

# Hard ceiling on concurrency for the spawned v2dl process. Refuses to
# go higher even if the user passes a bigger ``--max-worker`` - the goal
# of this wrapper is *polite* archival, not maximum throughput.
MAX_WORKER_CEILING = 3
RATE_LIMIT_CEILING_KBPS = 2000


# ----- argument parsing -----------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Local archival sync for V2PH-Downloader (personal use only).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    # ---- discover ----------------------------------------------------------
    pd = sub.add_parser(
        "discover",
        help="Enumerate top-level listing URLs into a watch file.",
    )
    pd.add_argument(
        "target",
        choices=["companies", "actors"],
        help=(
            "What to enumerate. 'companies' scrapes /company/ once and "
            "writes data/sync/companies.xlsx (url, name, total, "
            "是否采集). 'actors' paginates through every page of every "
            "company in companies.xlsx and writes data/sync/actors.xlsx "
            "with the same four columns (total = how many pages the "
            "actor was seen on across the walked companies)."
        ),
    )
    pd.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Override the output file. Defaults to data/sync/companies.xlsx "
            "(or data/sync/actors.xlsx for the actors target). Use a .txt "
            "extension to fall back to the plain-text format if you don't "
            "want the xlsx workflow."
        ),
    )
    pd.add_argument(
        "--companies-file",
        type=Path,
        default=DEFAULT_COMPANIES_FILE,
        help=(
            "When target=actors, the companies file to walk through. "
            "Accepts either an xlsx or a plain-text URL list. "
            "Default: data/sync/companies.xlsx."
        ),
    )
    pd.add_argument(
        "--only-selected",
        action="store_true",
        help=(
            "When target=actors, walk only companies with 是否采集 == 1 in "
            "the xlsx (much faster, obviously misses actors from unselected "
            "companies). Ignored for plain-text companies files."
        ),
    )
    pd.add_argument(
        "--max-pages-per-company",
        type=int,
        default=0,
        help=(
            "When target=actors, cap pagination depth per company. "
            "0 = walk every page (default). Set e.g. 3 to cap each "
            "company to the first 3 pages if a full crawl is too heavy."
        ),
    )

    # ---- sync --------------------------------------------------------------
    ps = sub.add_parser(
        "sync",
        help="Drive v2dl against the watch file.",
    )
    ps.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_COMPANIES_FILE,
        help=(
            "Watch-list file. xlsx is filtered by 是否采集 == 1, plain-text "
            "is consumed verbatim. Default: data/sync/companies.xlsx (falls "
            "back to data/sync/companies.txt if the xlsx is missing)."
        ),
    )
    ps.add_argument(
        "--destination",
        "-d",
        type=Path,
        required=True,
        help="Local archive root (passed through to v2dl -d).",
    )
    ps.add_argument(
        "--mode",
        choices=["full", "incremental"],
        default="incremental",
        help=(
            "incremental: only check page 1 of each listing - new albums "
            "appear first, downloaded albums get skipped via download_log. "
            "full: paginate through every listing page (use weekly at most)."
        ),
    )
    ps.add_argument(
        "--max-worker",
        type=int,
        default=2,
        help=f"Per-album download concurrency (capped at {MAX_WORKER_CEILING}).",
    )
    ps.add_argument(
        "--rate-limit",
        type=int,
        default=1000,
        help=f"Per-download speed limit in kbps (capped at {RATE_LIMIT_CEILING_KBPS}).",
    )
    ps.add_argument(
        "--force-download",
        action="store_true",
        help=(
            "Retry albums that are empty or partially downloaded (on-disk "
            "image count below the site's listed count). Albums already "
            "complete are still skipped; existing image files are not "
            "re-fetched."
        ),
    )
    ps.add_argument(
        "--dry-run",
        action="store_true",
        help="Print the planned v2dl invocation and exit without running it.",
    )

    return p


# ----- discover -------------------------------------------------------------

def _build_v2dl_namespace(stub_url: str) -> argparse.Namespace:
    """Build an args Namespace that satisfies V2DLApp.init().

    We piggy-back on v2dl's own CLI parser so we never fall behind when new
    fields get added.
    """
    from v2dl.cli import parse_arguments

    return parse_arguments([stub_url])


async def _scrape_html(url: str) -> str:
    """Drive the v2dl web bot through one page fetch and return the HTML.

    Convenience wrapper for callers that only need a single page (e.g.
    the /company/ index). For multi-page workflows (``discover actors``)
    use :class:`_BotSession` instead - spinning Chrome up per fetch costs
    ~5-10s each and adds up fast across 200 companies × N pages.
    """
    async with _BotSession() as session:
        return await session.fetch(url)


class _BotSession:
    """Async context manager that reuses one v2dl web bot for many fetches.

    The v2dl web bot is expensive to start (it boots Chromium via
    DrissionPage / Selenium and runs the Cloudflare clearance dance).
    For workflows that walk hundreds of pages - like paginating actors
    across every company - we boot it once at the start and tear it
    down at the end, instead of paying that cost per request.

    Use as::

        async with _BotSession() as session:
            for url in urls:
                html_content = await session.fetch(url)
    """

    def __init__(self) -> None:
        self.app: Any = None

    async def __aenter__(self) -> "_BotSession":
        from v2dl import V2DLApp

        # Stub URL just satisfies v2dl's argparser; every actual fetch
        # passes its own URL into ``auto_page_scroll`` directly.
        args = _build_v2dl_namespace("https://www.v2ph.com/")
        args.terminate = True
        self.app = V2DLApp()
        await self.app.init(args)
        return self

    async def __aexit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        if self.app is not None:
            try:
                self.app.bot.close_driver()
            except Exception:
                pass

    async def fetch(self, url: str) -> str:
        if self.app is None:
            raise RuntimeError("_BotSession used outside of `async with` block")
        return await self.app.bot.auto_page_scroll(url, page_sleep=0)


def _extract_paths(html_content: str, kind: str) -> list[str]:
    """Pull /company/<name> or /actor/<name> hrefs out of an index page.

    The v2ph layout uses regular ``<a href="/company/Beautyleg">...`` links
    inside a flat list - no special class is reliably present across the
    different listing pages, so we lean on the URL prefix instead.

    Kept as the simple slug-only API; richer extraction (display name +
    count) lives in :func:`_extract_listing_entries`.
    """
    tree = html.fromstring(html_content)
    seen: set[str] = set()
    out: list[str] = []
    prefix = f"/{kind}/"
    pattern = re.compile(rf"^{re.escape(prefix)}([^/?#]+)$")
    for href in tree.xpath("//a/@href"):
        if not href:
            continue
        path = href.split("?", 1)[0].split("#", 1)[0]
        m = pattern.match(path)
        if not m:
            continue
        name = m.group(1)
        if name in seen:
            continue
        seen.add(name)
        out.append(name)
    return out


# Total-count patterns seen on v2ph listing cards. The site uses several
# CJK suffixes depending on the language toggle (套 / 套图 / 部 / 张) and
# occasionally English ("sets"). Each pattern captures the integer only.
_COUNT_RE = re.compile(
    r"(\d[\d,]*)\s*(?:套图|套|部|张|張|張圖|sets?)",
    re.IGNORECASE,
)


def _normalize_text(s: str) -> str:
    """Collapse whitespace runs in ``s`` to single spaces."""
    return re.sub(r"\s+", " ", s).strip()


def _element_text(el: html.HtmlElement) -> str:
    """Return all visible text descendants of ``el`` as one normalised string."""
    try:
        parts = [t for t in el.xpath(".//text()") if t and t.strip()]
    except Exception:
        return ""
    return _normalize_text(" ".join(parts))


def _find_card_ancestor(
    a: html.HtmlElement, prefix: str = "/company/"
) -> html.HtmlElement | None:
    """Walk up from ``a`` and return the largest ancestor that still only
    contains links pointing to the same ``prefix``-prefixed slug.

    The point is to widen the search scope just enough to catch a
    sibling element (e.g. ``<small>456 套</small>``) that belongs to the
    same card, without bleeding into the next card and stealing its
    count / name. Returns ``None`` when no useful ancestor exists (the
    immediate parent already mixes multiple companies / actors).
    """
    last_good: html.HtmlElement | None = None
    cur = a.getparent()
    xpath_expr = f'.//a[starts-with(@href, "{prefix}")]'
    for _ in range(6):  # don't walk forever; 6 levels covers any sane layout
        if cur is None:
            break
        try:
            links = cur.xpath(xpath_expr)
        except Exception:
            break
        slugs: set[str] = set()
        for link in links:
            href = (link.get("href") or "").split("?", 1)[0].split("#", 1)[0]
            slugs.add(href)
        if len(slugs) > 1:
            return last_good  # crossing into another card; stop
        last_good = cur
        cur = cur.getparent()
    return last_good


def _extract_listing_entries(
    tree: html.HtmlElement, kind: str
) -> list[dict[str, str]]:
    """Pull (url, slug, name, total) tuples for every ``/{kind}/<slug>`` link
    on the page.

    Strategy:

    1. Find each ``<a href="/{kind}/<slug>">``.
    2. Read display name + (for ``kind=company``) count from the link's
       own text first - v2ph inlines both inside the clickable card.
    3. If the link is image-only, widen the search via
       :func:`_find_card_ancestor` - but never to a scope that contains
       a different ``/{kind}/<other>``, so we can't steal a neighbour's
       count.
    4. Actor cards don't normally carry a count (``套`` on album cards
       refers to album page count, not the actor's total albums), so we
       skip the count regex for ``kind=actor``; ``total`` is populated
       later by counting appearances across pages.

    Falls back to slug when the page layout has no inline name. Failure
    modes stay non-fatal so a single weird card can't blow up the run.
    """
    prefix = f"/{kind}/"
    pattern = re.compile(rf"^/{re.escape(kind)}/([^/?#]+)$")
    entries: list[dict[str, str]] = []
    seen: set[str] = set()

    parse_count = (kind == "company")

    xpath_expr = f'//a[starts-with(@href, "{prefix}")]'
    for a in tree.xpath(xpath_expr):
        href = (a.get("href") or "").split("?", 1)[0].split("#", 1)[0]
        m = pattern.match(href)
        if not m:
            continue
        slug = m.group(1)
        if slug in seen:
            continue
        seen.add(slug)

        link_text = _element_text(a)
        total = ""
        name = ""

        count_match = _COUNT_RE.search(link_text) if parse_count else None
        if count_match:
            total = count_match.group(1).replace(",", "")
            name = _normalize_text(
                link_text[: count_match.start()] + link_text[count_match.end() :]
            )
        else:
            name = link_text

        # Wider search via card ancestor (only relevant when the link
        # itself is image-only and we still need the name / count).
        need_count_search = parse_count and not total
        if not name or need_count_search:
            card = _find_card_ancestor(a, prefix=prefix)
            if card is not None:
                card_text = _element_text(card)
                if need_count_search:
                    m2 = _COUNT_RE.search(card_text)
                    if m2:
                        total = m2.group(1).replace(",", "")
                        if not name:
                            name = _normalize_text(
                                card_text[: m2.start()] + card_text[m2.end() :]
                            )
                if not name:
                    name = card_text

        if not name:
            # No visible label anywhere - probably an image-only card.
            # Fall back to the URL slug so the user still has something
            # readable in Excel.
            name = slug

        if len(name) > 80:
            name = name[:80].rstrip() + "…"

        entries.append(
            {
                "url": f"https://www.v2ph.com/{kind}/{slug}?hl=zh-Hans",
                "slug": slug,
                "name": name,
                "total": total,
            }
        )
    return entries


def _extract_company_entries(html_content: str) -> list[dict[str, str]]:
    """Thin wrapper kept for backwards compatibility / readability."""
    tree = html.fromstring(html_content)
    return _extract_listing_entries(tree, kind="company")


def _write_url_list(output: Path, kind: str, names: list[str]) -> None:
    """Write a plain-text watch list (one URL per line). Used for actors."""
    output.parent.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with output.open("w", encoding="utf-8") as f:
        f.write(f"# v2dl sync watch list ({kind}s)\n")
        f.write(f"# Generated: {timestamp}\n")
        f.write(f"# Count: {len(names)}\n")
        f.write("# One URL per line, '#' = comment, processed lines auto-prefixed with '# '\n")
        for name in names:
            f.write(f"https://www.v2ph.com/{kind}/{name}?hl=zh-Hans\n")


def _require_openpyxl() -> Any:
    """Import ``openpyxl`` lazily so the script still runs for non-xlsx flows
    on a fresh install that hasn't pulled the optional dep yet."""
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required for the xlsx workflow but is not installed. "
            "Run `pip install openpyxl` (it's already a transitive dep of "
            "DrissionPage's datarecorder, so a `pip install -r requirements.txt` "
            "in the project venv will pick it up)."
        ) from e
    return openpyxl


def _load_existing_collect_marks(path: Path) -> dict[str, str]:
    """Read the existing xlsx and return ``{url: 是否采集}`` so re-running
    ``discover companies`` doesn't wipe the user's hand-edited selections.

    Anything other than a clean ``0`` / ``1`` is preserved verbatim (e.g.
    ``"1"`` vs ``1`` vs blank cell) so the user's spreadsheet stays
    byte-identical on the columns they care about.
    """
    if not path.exists():
        return {}
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[discover] WARN: failed to read existing {path} ({e});"
              " any prior 是否采集 marks will be lost.", file=sys.stderr)
        return {}
    marks: dict[str, str] = {}
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return {}
        # Locate columns by name so the layout can shift defensively.
        try:
            url_idx = header.index("url")
        except ValueError:
            return {}
        collect_idx = None
        for i, h in enumerate(header):
            if h and isinstance(h, str) and h.startswith("是否采集"):
                collect_idx = i
                break
        if collect_idx is None:
            return {}
        for row in rows:
            if not row or url_idx >= len(row):
                continue
            url = row[url_idx]
            if not url:
                continue
            mark = row[collect_idx] if collect_idx < len(row) else None
            if mark is None:
                continue
            marks[str(url).strip()] = str(mark).strip()
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return marks


def _write_listing_xlsx(
    output: Path,
    entries: list[dict[str, str]],
    existing_marks: dict[str, str],
    sheet_title: str,
) -> None:
    """Generic xlsx writer used by both ``discover companies`` and
    ``discover actors``.

    Writes the four required columns (``url`` / ``name`` / ``total`` /
    ``是否采集(0/1)``), preserves any ``是否采集`` value from
    ``existing_marks`` so re-running discover doesn't wipe user edits,
    and applies a small bit of styling (bold/frozen header, autofilter).
    """
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    output.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(XLSX_HEADER)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFE2E2E2")
    for col_idx, _ in enumerate(XLSX_HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    preserved = 0
    for entry in entries:
        url = entry["url"]
        mark = existing_marks.get(url, "0")
        if mark and mark != "0":
            preserved += 1
        try:
            mark_value: Any = int(mark)
        except (TypeError, ValueError):
            mark_value = mark or 0
        try:
            total_value: Any = int(entry["total"]) if entry["total"] else ""
        except ValueError:
            total_value = entry["total"]
        ws.append([url, entry["name"], total_value, mark_value])

    column_widths = {1: 56, 2: 28, 3: 10, 4: 16}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(XLSX_HEADER))
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    wb.save(output)
    if preserved:
        print(f"[discover] preserved 是否采集 marks on {preserved} existing rows")


def _write_companies_xlsx(
    output: Path,
    entries: list[dict[str, str]],
    existing_marks: dict[str, str],
) -> None:
    """Render company ``entries`` into ``output`` as an xlsx."""
    _write_listing_xlsx(output, entries, existing_marks, sheet_title="companies")


def _write_actors_xlsx(
    output: Path,
    entries: list[dict[str, str]],
    existing_marks: dict[str, str],
) -> None:
    """Render actor ``entries`` into ``output`` as an xlsx."""
    _write_listing_xlsx(output, entries, existing_marks, sheet_title="actors")


async def _discover_companies(output: Path) -> int:
    url = "https://www.v2ph.com/company/?hl=zh-Hans"
    print(f"[discover] fetching {url}")
    html_content = await _scrape_html(url)
    entries = _extract_company_entries(html_content)
    if not entries:
        raise SystemExit(
            "No /company/<name> links found on the index page. Either the layout "
            "changed (update the XPath in this script) or the request was blocked "
            "before reaching the real page."
        )

    if output.suffix.lower() != ".xlsx":
        # Allow ``--output companies.txt`` for the legacy plain-text flow,
        # but the default path is .xlsx.
        names = [e["slug"] for e in entries]
        _write_url_list(output, "company", names)
        print(f"[discover] wrote {len(entries)} companies to {output} (plain text)")
        return len(entries)

    existing_marks = _load_existing_collect_marks(output)
    _write_companies_xlsx(output, entries, existing_marks)
    print(
        f"[discover] wrote {len(entries)} companies to {output}\n"
        f"           open it in Excel and set 是否采集 = 1 for the ones "
        "you want to sync, then save."
    )
    return len(entries)


async def _paginate_company_for_actors(
    session: _BotSession,
    company_url: str,
    *,
    max_pages: int = 0,
) -> dict[str, dict[str, str]]:
    """Walk every page of a single company listing and collect actor links.

    Returns ``{slug: {url, slug, name, total}}`` where ``total`` is the
    number of pages on this company that link to the actor at least once
    (dedup-within-page, count-across-pages). Useful as a popularity
    indicator in the spreadsheet.

    Stops at ``max_pages`` if non-zero, or when v2dl's pagination probe
    says we're past the last page, or when a page fetch fails.
    """
    from v2dl.scraper.tools import UrlHandler

    results: dict[str, dict[str, str]] = {}
    page = 1
    max_page_total: int | None = None

    while True:
        page_url = UrlHandler.add_page_num(company_url, page)
        if max_page_total is not None:
            print(f"[discover]     page {page}/{max_page_total}")
        else:
            print(f"[discover]     page {page}")
        try:
            html_content = await session.fetch(page_url)
        except Exception as e:
            print(
                f"[discover]     fetch error on page {page}: {e}",
                file=sys.stderr,
            )
            break
        if not html_content or "Failed" in html_content:
            print(
                f"[discover]     page {page} fetch failed, stopping",
                file=sys.stderr,
            )
            break
        try:
            tree = html.fromstring(html_content)
        except Exception as e:
            print(
                f"[discover]     page {page} parse error: {e}",
                file=sys.stderr,
            )
            break

        page_entries = _extract_listing_entries(tree, kind="actor")
        for entry in page_entries:
            slug = entry["slug"]
            if slug in results:
                # Bump the page-appearance count for this actor.
                try:
                    cur = int(results[slug]["total"] or 0)
                except ValueError:
                    cur = 0
                results[slug]["total"] = str(cur + 1)
            else:
                e = dict(entry)
                e["total"] = "1"
                results[slug] = e

        if max_page_total is None:
            try:
                max_page_total = max(UrlHandler.get_max_page(tree) or 1, 1)
            except Exception:
                max_page_total = 1

        if not page_entries and page > 1:
            # Defensive: a page past the last sometimes returns 200 with
            # no album cards. Stop instead of looping forever.
            break
        if max_pages and page >= max_pages:
            break
        if page >= max_page_total:
            break
        page += 1

    return results


async def _discover_actors(
    companies_file: Path,
    output: Path,
    *,
    only_selected: bool = False,
    max_pages_per_company: int = 0,
) -> int:
    """Enumerate every actor across every walked company page.

    ``only_selected`` restricts the walk to companies with 是否采集 == 1
    in the xlsx (much faster, but obviously misses actors from companies
    you haven't subscribed to). ``max_pages_per_company`` caps the
    pagination depth per company - 0 means walk every page.
    """
    if not companies_file.exists():
        if (
            companies_file == DEFAULT_COMPANIES_FILE
            and LEGACY_COMPANIES_TXT.exists()
        ):
            print(
                f"[discover] {companies_file} not found, falling back to legacy "
                f"{LEGACY_COMPANIES_TXT}"
            )
            companies_file = LEGACY_COMPANIES_TXT
        else:
            raise SystemExit(
                f"Companies file not found: {companies_file}. "
                "Run 'discover companies' first."
            )

    company_pairs: list[tuple[str, str]] = []
    if companies_file.suffix.lower() == ".xlsx":
        if only_selected:
            company_pairs = _load_urls_from_xlsx(companies_file)
            label = "selected (是否采集 == 1)"
        else:
            company_pairs = _load_urls_from_xlsx_all(companies_file)
            label = "all"
    else:
        for line in companies_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            company_pairs.append((line, ""))
        label = "all"

    if not company_pairs:
        raise SystemExit(
            f"No companies to walk from {companies_file} "
            f"(filter: {label}). "
            "Did you forget to mark 是否采集 = 1?"
        )

    print(
        f"[discover] walking {len(company_pairs)} {label} companies for actors"
    )

    all_actors: dict[str, dict[str, str]] = {}

    # One Chrome instance for the whole crawl - per-fetch boot would
    # otherwise dominate the runtime.
    async with _BotSession() as session:
        for i, (curl, cname) in enumerate(company_pairs, 1):
            label_text = cname or curl
            print(
                f"[discover] ({i}/{len(company_pairs)}) {label_text}"
            )
            try:
                company_actors = await _paginate_company_for_actors(
                    session, curl, max_pages=max_pages_per_company
                )
            except Exception as e:
                print(
                    f"[discover]   company walk failed: {e}; skipping",
                    file=sys.stderr,
                )
                continue
            added = 0
            for slug, entry in company_actors.items():
                if slug in all_actors:
                    try:
                        cur = int(all_actors[slug]["total"] or 0)
                        inc = int(entry["total"] or 0)
                    except ValueError:
                        cur, inc = 0, 0
                    all_actors[slug]["total"] = str(cur + inc)
                else:
                    all_actors[slug] = dict(entry)
                    added += 1
            print(
                f"[discover]   +{added} new actors "
                f"(running total {len(all_actors)})"
            )
            # Polite pause between companies. We're inside a long crawl,
            # this gives the site (and us) a breather.
            time.sleep(INTER_URL_SLEEP_SECONDS)

    entries = list(all_actors.values())
    # Sort by total descending for a more useful initial view in Excel.
    def _sort_key(e: dict[str, str]) -> tuple[int, str]:
        try:
            t = -int(e["total"] or 0)
        except ValueError:
            t = 0
        return (t, e["name"])
    entries.sort(key=_sort_key)

    if output.suffix.lower() == ".xlsx":
        existing_marks = _load_existing_collect_marks(output)
        _write_actors_xlsx(output, entries, existing_marks)
        print(
            f"[discover] wrote {len(entries)} actors to {output}\n"
            f"           open it in Excel and set 是否采集 = 1 for the actors "
            "you want to sync, then save."
        )
    else:
        # Plain-text fallback (no count info), preserved for users who
        # passed ``--output actors.txt``.
        _write_url_list(output, "actor", [e["slug"] for e in entries])
        print(f"[discover] wrote {len(entries)} actors to {output} (plain text)")

    return len(entries)


# ----- sync -----------------------------------------------------------------

def _load_urls_from_xlsx_all(path: Path) -> list[tuple[str, str]]:
    """Return every (url, name) row from the xlsx, regardless of 是否采集.

    Used by ``discover actors`` which legitimately wants to walk every
    company (the user may not have decided which ones to mirror yet).
    """
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, str]] = []
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        try:
            url_idx = header.index("url")
        except ValueError:
            return []
        try:
            name_idx = header.index("name")
        except ValueError:
            name_idx = None
        for row in rows:
            if not row or url_idx >= len(row):
                continue
            url = row[url_idx]
            if not url:
                continue
            name = ""
            if name_idx is not None and name_idx < len(row) and row[name_idx]:
                name = str(row[name_idx]).strip()
            out.append((str(url).strip(), name))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def _load_urls_from_xlsx(path: Path) -> list[tuple[str, str]]:
    """Return ``[(url, name), ...]`` for rows with 是否采集 == 1.

    Rows that don't parse cleanly are silently skipped (with a warning) so
    one bad cell can't block the whole sync run.
    """
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    selected: list[tuple[str, str]] = []
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        try:
            url_idx = header.index("url")
        except ValueError:
            raise SystemExit(
                f"{path}: missing required column 'url'. Re-run "
                "`discover companies` to regenerate the sheet."
            )
        try:
            name_idx = header.index("name")
        except ValueError:
            name_idx = None
        collect_idx = None
        for i, h in enumerate(header):
            if h and isinstance(h, str) and h.startswith("是否采集"):
                collect_idx = i
                break
        if collect_idx is None:
            raise SystemExit(
                f"{path}: missing required column '是否采集(0/1)'. Re-run "
                "`discover companies` to regenerate the sheet."
            )
        for row_idx, row in enumerate(rows, start=2):
            if not row:
                continue
            if url_idx >= len(row) or collect_idx >= len(row):
                continue
            url = row[url_idx]
            collect = row[collect_idx]
            if collect is None:
                continue
            # Excel cells come back as int or str depending on whether
            # the user typed "1" or 1. Normalise.
            try:
                collect_int = int(str(collect).strip())
            except ValueError:
                print(
                    f"[sync] row {row_idx}: unrecognised 是否采集 value "
                    f"{collect!r}, skipping",
                    file=sys.stderr,
                )
                continue
            if collect_int != 1:
                continue
            if not url:
                continue
            name = ""
            if name_idx is not None and name_idx < len(row) and row[name_idx]:
                name = str(row[name_idx]).strip()
            selected.append((str(url).strip(), name))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return selected


def _materialize_xlsx_to_tempfile(
    xlsx_path: Path, selected: list[tuple[str, str]]
) -> Path:
    """Write ``selected`` URLs to a temp .txt v2dl can consume via
    ``--input-file``. Returns the temp file's path so the caller can
    clean it up afterwards.
    """
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    fd, name = tempfile.mkstemp(prefix="v2dl-sync-", suffix=".txt", text=True)
    tmp_path = Path(name)
    try:
        with open(fd, "w", encoding="utf-8") as f:
            f.write(f"# Generated from {xlsx_path} at {timestamp}\n")
            f.write(f"# Selected rows (是否采集 == 1): {len(selected)}\n")
            for url, label in selected:
                if label:
                    f.write(f"# {label}\n")
                f.write(url + "\n")
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
    return tmp_path


def _resolve_sync_input(input_path: Path) -> tuple[Path, Path | None, int]:
    """Return ``(file_to_feed_v2dl, temp_file_to_cleanup, selected_count)``.

    - For .xlsx input, filters by 是否采集 == 1 and materialises a temp
      .txt for v2dl. ``temp_file_to_cleanup`` is the path the caller
      should ``unlink`` after v2dl exits.
    - For .txt input, just returns the path as-is (no cleanup needed).
    - When the default xlsx is missing but a legacy companies.txt exists,
      transparently use that so existing installs keep working.
    """
    if not input_path.exists():
        if (
            input_path == DEFAULT_COMPANIES_FILE
            and LEGACY_COMPANIES_TXT.exists()
        ):
            print(
                f"[sync] {input_path} not found, falling back to legacy "
                f"{LEGACY_COMPANIES_TXT}"
            )
            input_path = LEGACY_COMPANIES_TXT
        else:
            raise SystemExit(
                f"[sync] watch file not found: {input_path}\n"
                "[sync] run `python scripts/sync_local.py discover companies` first."
            )

    if input_path.suffix.lower() != ".xlsx":
        return input_path, None, -1

    selected = _load_urls_from_xlsx(input_path)
    if not selected:
        raise SystemExit(
            f"[sync] no rows with 是否采集 == 1 in {input_path}.\n"
            "[sync] open the file in Excel, set 是否采集 = 1 for the companies "
            "you want to sync, save, and re-run."
        )
    tmp_path = _materialize_xlsx_to_tempfile(input_path, selected)
    print(
        f"[sync] {len(selected)} companies selected from {input_path.name} "
        f"(temp watch file: {tmp_path})"
    )
    return tmp_path, tmp_path, len(selected)


def _sync(args: argparse.Namespace) -> int:
    try:
        input_for_v2dl, cleanup_path, _selected = _resolve_sync_input(args.input)
    except SystemExit as e:
        print(str(e), file=sys.stderr)
        return 1

    try:
        max_worker = min(args.max_worker, MAX_WORKER_CEILING)
        rate_limit = min(args.rate_limit, RATE_LIMIT_CEILING_KBPS)
        if max_worker != args.max_worker:
            print(f"[sync] capping --max-worker {args.max_worker} -> {max_worker}")
        if rate_limit != args.rate_limit:
            print(f"[sync] capping --rate-limit {args.rate_limit} -> {rate_limit}")

        cmd: list[str] = [
            sys.executable,
            "-m",
            "v2dl",
            "--input-file",
            str(input_for_v2dl),
            "--destination",
            str(args.destination),
            "--max-worker",
            str(max_worker),
            "--rate-limit",
            str(rate_limit),
        ]
        if args.mode == "incremental":
            # Only look at the first page of each top-level listing. New
            # albums are listed first chronologically, and already-downloaded
            # albums get skipped instantly via download_log_path.
            cmd.extend(["--range", "1"])
        if args.force_download:
            cmd.append("--retry-incomplete")

        print("[sync] running:", " ".join(f'"{a}"' if " " in a else a for a in cmd))
        if args.dry_run:
            print("[sync] --dry-run set, exiting without executing")
            return 0

        try:
            return subprocess.call(cmd)
        except KeyboardInterrupt:
            print("\n[sync] interrupted by user", file=sys.stderr)
            return 130
    finally:
        if cleanup_path is not None:
            try:
                cleanup_path.unlink()
            except Exception:
                pass


# ----- entry point ----------------------------------------------------------

def main() -> int:
    parser = _build_parser()
    args = parser.parse_args()

    if args.cmd == "discover":
        if args.target == "companies":
            output = args.output or DEFAULT_COMPANIES_FILE
            asyncio.run(_discover_companies(output))
            return 0
        elif args.target == "actors":
            output = args.output or DEFAULT_ACTORS_FILE
            asyncio.run(
                _discover_actors(
                    args.companies_file,
                    output,
                    only_selected=args.only_selected,
                    max_pages_per_company=args.max_pages_per_company,
                )
            )
            return 0

    if args.cmd == "sync":
        return _sync(args)

    parser.error(f"unknown command: {args.cmd}")
    return 2


if __name__ == "__main__":
    sys.exit(main())
